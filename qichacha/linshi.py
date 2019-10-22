# -*- coding: utf-8 -*-
"""
Created on Thu Oct 17 10:08:53 2019

@author: h
"""
import requests
import lxml
import sys
from bs4 import BeautifulSoup
import xlwt
import time
import urllib
from ceshiqichacha import Craw,Craw_inside
from openpyxl import load_workbook
from openpyxl import Workbook
import signal

#打开表格，获取全部数据为rows列表，获取查询关键字列的数据，为data
wb = load_workbook('d:\\testone\\qichacha\\abc.xlsx')
ws=wb.active
rows=[]

for row in ws.iter_rows():
    rows.append(row)
    
def signal_handler(signal,frame):
    print('You pressed Ctrl+C!')

signal.signal(signal.SIGINT,signal_handler)


all_list=[]
   
#rou=[]
#rou=[('adasd','asddad'),'asadasd','b','c','d']
#for index,values in enumerate(rou):
#    print(str(index)+str(values[0]))
cc=1
for row in rows[201:1000] :
    key_word = str(row[0].value)
    print(str(cc)+str(key_word))
    cc=cc+1
    if  key_word:
        key_word2 = urllib.parse.quote(key_word)    
        url = r'https://www.qichacha.com/search?key='+key_word2
        result = Craw(url,key_word2)
        new_row =[]

        for m in range(len(row)):
            new_row.append(row[m].value)
        if  result :
            new_row.extend(result)
            all_list.append((new_row))
    
    while len(all_list)==50:
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.append(['公司抬头','姓名','手机号','邮箱','座机','QQ','微信','职位','部门','行业类型','客户类型','名单来源',
        '回访人','地址','是否注册','注册时间','注册账号','绑定手机','绑定邮箱','客服代表','经营范围',
        '公司','标签','法人','注册资本','成立日期','邮件','电话','地址','状态',
        '名称','电话','更多电话','官网','邮箱','法人','注册资本','实缴资本','状态',
        '成立时间','企业类型','所属行业','批准机关','区域','参保人数','人员规模','经营范围','简介','风险'
        ])
        for i in all_list:
            ws2.append(i)
        filename = time.strftime("%H%M%S", time.localtime())

        wb2.save('{}.xlsx'.format(filename))
        all_list=[]


#akk =[]
#for row in rows[:2] :
#    for m in range(len(row)) :
#        akk.append(row[m].value)
#    print(akk)    
#print(mingcheng)
#    wb.save("e:\\sample.xlsx")
#data = []   
#for i in range(0,len(rows)) :
#    data.append(rows[i][0].value)
#for m in range(len(rows)):
#    for n in range(len(m)):
#        rows[m]
#print(rows[1][:].value)   #所有行
#print (rows[0]) #获取第一行
#print (rows[0][0]) #获取第一行第一列的单元格对象
#print (rows[0][0].value) #获取第一行第一列的单元格对象的值
# 
#print (rows[len(rows)-1]) #获取最后行 print rows[-1]
#print (rows[len(rows)-1][len(rows[0])-1]) #获取第后一行和最后一列的单元格对象
#print (rows[len(rows)-1][len(rows[0])-1].value) #获取第后一行和最后一列的单元格对象的值