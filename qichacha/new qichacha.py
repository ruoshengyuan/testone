# -*- coding: utf-8 -*-
"""
Created on Thu Oct 17 10:08:53 2019

@author: h
"""
import requests
import lxml
import sys
from bs4 import BeautifulSoup
import xlwt
import time
import urllib
from new_qichacha_sousuo import Craw,Craw_inside
from openpyxl import load_workbook
from openpyxl import Workbook
import signal

#打开表格，获取全部数据为rows列表，获取查询关键字列的数据，为data
wb = load_workbook('d:\\testone\\qichacha\\1029北京天津制造业.xlsx')
ws=wb.active
rows=[]

for row in ws.iter_rows():
    rows.append(row)
    
def signal_handler(signal,frame):
    print('You pressed Ctrl+C!')

signal.signal(signal.SIGINT,signal_handler)


all_list=[]
   
#rou=[]
#rou=[('adasd','asddad'),'asadasd','b','c','d']
#for index,values in enumerate(rou):
#    print(str(index)+str(values[0]))
cc=1
for row in rows[201:] :
    key_word = str(row[1].value)
    print(str(cc)+str(key_word))
    cc=cc+1
    if  key_word:
        key_word2 = urllib.parse.quote(key_word)    
        url = r'https://www.qichacha.com/search?key='+key_word2
        result = Craw(url,key_word2)
        new_row =[]

        for m in range(len(row)):
            new_row.append(row[m].value)
        if  result :
            new_row.extend(result)
            all_list.append((new_row))
    
    while len(all_list)==100:
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.append(['序号',	'企业名称',	'经营状态',	'法定代表人',	
                    '注册资本',	'成立日期',	'所属省份',	'所属城市',	
                    '电话',	'更多电话',	'邮箱',	'统一社会信用代码',	
                    '纳税人识别号',	'注册号',	'组织机构代码',	'参保人數',	
                    '企业类型',	'所属行业',	'网址',	'企业地址',	'经营范围','公司','标签','实缴资本','人员规模','简介','风险'
        ])
        for i in all_list:
            ws2.append(i)
        filename = time.strftime("%H%M%S", time.localtime())

        wb2.save('{}.xlsx'.format(filename))
        all_list=[]


#akk =[]
#for row in rows[:2] :
#    for m in range(len(row)) :
#        akk.append(row[m].value)
#    print(akk)    
#print(mingcheng)
#    wb.save("e:\\sample.xlsx")
#data = []   
#for i in range(0,len(rows)) :
#    data.append(rows[i][0].value)
#for m in range(len(rows)):
#    for n in range(len(m)):
#        rows[m]
#print(rows[1][:].value)   #所有行
#print (rows[0]) #获取第一行
#print (rows[0][0]) #获取第一行第一列的单元格对象
#print (rows[0][0].value) #获取第一行第一列的单元格对象的值
# 
#print (rows[len(rows)-1]) #获取最后行 print rows[-1]
#print (rows[len(rows)-1][len(rows[0])-1]) #获取第后一行和最后一列的单元格对象
#print (rows[len(rows)-1][len(rows[0])-1].value) #获取第后一行和最后一列的单元格对象的值