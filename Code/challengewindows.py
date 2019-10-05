from openpyxl import load_workbook
from openpyxl import Workbook 
from datetime import datetime

path = 'D:\\我的文档\\GitHub\\testone\\Code\\courses.xlsx'
    
def combine():
    wb = load_workbook(path)
    # ws = wb.active #dangqianbiao
    # wb.remove("combine")
    for sheet in wb :
        if sheet.title == 'combine' :
            del wb['combine']
    ws3 = wb.create_sheet("combine")

    ws1 = wb["students"]
    ws2 = wb["time"]
    #获得全部表1数据到表3
    for i,row in enumerate(ws1.iter_rows()):
        for j,cell in enumerate(row):
            ws3.cell(row=i+1,column=j+1,value=cell.value)

    # 表3，表2的数据进行关联，知道表三的行数有多少，逐行关联从行号1开始


    for row3 in range(1,ws3.max_row+1):
        for row2 in range(1,ws2.max_row+1):
            if ws3.cell(row=row3,column=2).value==ws2.cell(row=row2,column=2).value :
                ws3.cell(row=row3,column=4).value=ws2.cell(row=row2,column=3).value 

    # 打印所有表的名称
    for sheet in wb :
        print (sheet.title)

    wb.save(path)


def split():
    wb = load_workbook(path)
    ws3 = wb["combine"]
    ws3.cell(row=1,column=5).value ='年份'
    mmm = []

    for row3 in range(2,ws3.max_row+1):
        kkk=datetime.strftime(ws3.cell(row=row3,column=1).value,'%Y')
        ws3.cell(row=row3,column=5).value =kkk
        mmm.append(kkk)



    nnn=list(set(mmm))
    
    for i in nnn :
        path_new=i+'.xlsx'
        wb_new = Workbook()
        ws4 = wb_new.active
        linshi =[]
        ooo=()
        for row3 in range(2,ws3.max_row+1):
            if i==ws3.cell(row=row3,column=5).value :
                ooo = (ws3.cell(row=row3,column=1).value,ws3.cell(row=row3,column=2).value,ws3.cell(row=row3,column=3).value,ws3.cell(row=row3,column=4).value)
                linshi.append(ooo)
        row2 =1
        for i in linshi :
            column2 =1
            for k in i :
                ws4.cell(row=row2,column=column2).value=str(k)

                column2=1+column2
            row2=1+row2
        wb_new.save(path_new)

    wb.save(path)




if __name__ == '__main__':
    combine()
    split()
    