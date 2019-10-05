from openpyxl import load_workbook
from openpyxl import Workbook 
import datetime

def combine():
	wb = load_workbook('/home/shiyanlou/testone/Code/courses.xlsx')
	ws = wb.active #dangqianbiao
	ws3 = wb.create_sheet("combine")
	for sheet in wb :
		print (sheet.title)
	ws1 = wb["students"]
	ws2 = wb["time"]
	for i,row in enumerate(ws1.iter_rows()):
    	for j,cell in enumerate(row):
        	ws3.cell(row=i+1,column=j+1,value=cell.value)
    
	# for row in ws3.rows :
	# 	print(row[0].value)
		# for cell in row :
		# 	print(cell.value)
def split():
	pass


if __name__ == '__main__':
	combine()
	split()
	